import os
from datetime import date

import requests
import xlwt
import openpyxl as openpyxl
import yaml

from Class.db import DB


class SYSTEM():
    """Work with system commands"""

    def __init__(self):
        """Initsialition"""
        self.db = DB()

    def seperate(self, list):

        columns = []
        colNumber = 1
        startRow = 2
        insert = {}

        while list.cell(row=1, column=colNumber).value is not None:
            columns.append(list.cell(row=1, column=colNumber).value)
            colNumber += 1

        while list.cell(row=startRow, column=1).value is not None:

            insert[startRow-2] = {}
            for ind in range(len(columns)):
                cell = list.cell(row=startRow, column=ind + 1).value
                if cell is not None:
                    if isinstance(cell, date):
                        insert[startRow-2][columns[ind]] = int(cell.timestamp())
                    else:
                        insert[startRow-2][columns[ind]] = cell

            startRow += 1

        if len(insert)==0:
            insert = None

        return insert


    def importToTable(self, file):

        path = os.getcwd()
        db_path = os.path.join(path, file)

        wb = openpyxl.load_workbook(f'{db_path}')

        table = file.split(".")
        table = table[0].split(' (')
        table = table[0].split('_')

        forID = table[1][:len(table[1]) - 1]
        insertDB = DB()

        sheetnames = wb.sheetnames

        if 'Добавить' in sheetnames:

            addedList = wb["Добавить"]

            results = self.seperate(list=addedList)

            if results is not None:
                for ind in results:
                    cols = []
                    vals = []
                    for key,value in results[ind].items():
                        cols.append(key)
                        vals.append(value)

                    insertDB.insert(
                        table=table[1],
                        columns=cols,
                        values=vals,
                        closed=False
                    )


        if 'Редактировать' in sheetnames:

            addedList = wb["Редактировать"]
            sets = {}
            conditions = {}

            results = self.seperate(list=addedList)

            if results is not None:
                for ind in results:
                    for key,value in results[ind].items():
                        if key != f"{forID}_id":
                            sets[key] = value
                        else:
                            conditions[f"{forID}_id"]=value

                    insertDB.update(
                        table=table[1],
                        sets=sets,
                        conditions=conditions,
                        closed=False
                    )


        if 'Удалить' in sheetnames:

            addedList = wb["Удалить"]

            results = self.seperate(list=addedList)

            sets = {}

            if results is not None:

                for ind in results:
                    for key,value in results[ind].items():
                        sets[key] = value

                    insertDB.deleteRow(
                        table=table[1],
                        conditions=sets,
                        closed=False
                    )

        if 'Локализация' in sheetnames:

            addedList = wb["Локализация"]

            results = self.seperate(list=addedList)

            if results is not None:

                for ind in results:

                    checked = self.db.fetchone(
                        table="localize",
                        conditions={"alias": results[ind]['alias']},
                        closed=False
                    )

                    if checked is None:
                        cols = []
                        vals = []
                        for key, value in results[ind].items():
                            cols.append(key)
                            vals.append(value)

                        insertDB.insert(
                            table=table[1],
                            columns=cols,
                            values=vals,
                            closed=False
                        )

                    else:
                        sets = {}
                        conditions = {}
                        for key, value in results[ind].items():
                            sets[key] = value
                            if key == "alias":
                                conditions[f"alias"] = value

                        insertDB.update(
                            table=table[1],
                            sets=sets,
                            conditions=conditions,
                            closed=False
                        )


        os.remove(f'{db_path}')


    def getAgents(self):

        path = os.getcwd()
        db_path = os.path.join(path, "config.yaml")

        with open(db_path, 'r', encoding='utf-8') as file:
            config = yaml.safe_load(file)
            authorized_ids = config['authorized_ids']

        res = self.db.fetchall(table='agents',closed=False)

        for agent in res:
            authorized_ids[agent[0]] = agent[2]

        return authorized_ids


    def getLang(self, user_id):

        user = self.db.fetchone(table='users', columns=['lang'], conditions={"user_id": user_id},closed=False)

        if user is None:
            lang = 'ru'
        else:
            lang = user[0]

        return lang


    def getEventDescLang(self, user_id):

        if self.getLang(user_id=user_id) == 'ru':
            eventDescLang = ''
        elif self.getLang(user_id=user_id) == 'uz':
            eventDescLang = '_uz'
        else:
            eventDescLang = ''

        return eventDescLang


    def getlocalize(self,user_id,alias):

        user = self.db.fetchone(
            table='users',
            columns=['lang','user_name'],
            conditions={"user_id": user_id},
            closed=False
        )
        if user is None:
            language = 'ru'
            userName = ''
        else:
            language = user[0]
            userName = user[1]

        text = self.db.fetchone(
            table='localize',
            columns=[f"{language}"],
            conditions={"alias": alias},
            closed=False
        )

        return text[0].replace("{user}",userName)


    def getReglocalize(self,user_id,alias):

        user = self.db.fetchone(
            table='temp_users',
            columns=['lang'],
            conditions={"user_id": user_id},
            closed=False
        )

        if user is None:
            user = ['ru']

        text = self.db.fetchone(
            table='localize',
            columns=[f"{user[0].lower()}"],
            conditions={"alias": alias},
            closed=False
        )

        return text[0]


    def getSubscribers(self,event_id):

        subcribers = self.db.fetchall(
            table='orders',
            columns=['user_id'],
            conditions={"event_id": event_id},
            closed=False
        )

        condition = list()

        for user in subcribers:

            condition.append(user[0])

        condition = tuple(condition)

        conditions = {}

        if len(condition)>1:
            conditions['user_id'] = {}

            conditions['user_id'][0] = condition
            conditions['user_id'][1] = ' in '
        else:
            conditions['user_id'] = condition[0]


        users = self.db.fetchall(
            table='users',
            columns=['user_name','phone','edu','lang'],
            conditions=conditions,
            closed=False
        )

        workbook = xlwt.Workbook(encoding='utf-8')

        mySheet = workbook.add_sheet("Подписчики")

        mySheet.write(0,0,'Имя')
        mySheet.write(0,1,'Немер телефона')
        mySheet.write(0,2,'Учебное место/Работа')
        mySheet.write(0,3,'Язык')

        row = 1

        for user in users:

            mySheet.write(row, 0, user[0])
            mySheet.write(row, 1, user[1])
            mySheet.write(row, 2, user[2])
            mySheet.write(row, 3, user[3].lower())
            row += 1

        workbook.save(f"scribers_id_{event_id}.xls")

        return f"scribers_id_{event_id}.xls"


    def getScribeCounts(self,event_id):
        subcribers = self.db.fetchall(
            table='orders',
            columns=['user_id'],
            conditions={"event_id": event_id},
            closed=False
        )

        return len(subcribers)


    def getToken(self):

        path = os.getcwd()
        db_path = os.path.join(path, "config.yaml")

        with open(db_path, 'r', encoding='utf-8') as file:
            config = yaml.safe_load(file)

        return config['TOKEN']